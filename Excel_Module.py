import openpyxl # for .xlsx
import os   # for Path
import datetime

# 데이터 읽어와 리스트에 넣기
def read_xlsx():
    # 엑셀파일 열기
    file = os.getcwd() + "/voca.xlsx"
    book = openpyxl.load_workbook(filename=file, read_only=False, data_only=False)

    # sheet1 불러오기
    sheet1 = book.worksheets[0]

    data = []
    for row in sheet1.rows:
        word = []
        for j in range(2): # 단어, 뜻
            word.append(row[j].value)
        if row[2].value is None:
            word.append(int(5)) # default 가중치
        else:
            word.append(int(row[2].value)) # 가중치 입력

        if row[3].value is None:
            word.append(datetime.datetime.now())
        else:
            word.append(row[3].value) # 날짜
        data.append(word)

    book.close()
    return data


# update 엑셀
def update_xlsx(data, max_rows):
    file = os.getcwd() + "/voca.xlsx"
    book = openpyxl.Workbook()

    for n in range(1, (len(data) + 1)):
        book.worksheets[0].cell(row=n, column=1).value = data[n - 1][0]
        book.worksheets[0].cell(row=n, column=2).value = data[n - 1][1]
        book.worksheets[0].cell(row=n, column=3).value = data[n - 1][2]
        book.worksheets[0].cell(row=n, column=4).value = data[n - 1][3] #datetime.datetime.strptime(str(data[n - 1][3]), '%Y-%m-%d').date()

    for n in range(len(data), len(data) - max_rows):
        book.worksheets[0].cell(row=n, column=1).value = ' '
        book.worksheets[0].cell(row=n, column=2).value = ' '
        book.worksheets[0].cell(row=n, column=3).value = ' '
        book.worksheets[0].cell(row=n, column=4).value = ' '

    book.save(file)
    book.close()
