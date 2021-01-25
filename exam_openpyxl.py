# -*- coding: utf-8 -*-

import datetime
import openpyxl
import os

sample_items = [(datetime.datetime.strptime('2021-01-25', '%Y-%m-%d'), 'Product Name 1', 5, 10000),
                (datetime.datetime.strptime('2021-01-25', '%Y-%m-%d'), 'Product Name 2', 10, 3000),
                (datetime.datetime.strptime('2021-01-25', '%Y-%m-%d'), 'Product Name 3', 3, 7000)]


def write_to_file(filepath):
    wb = openpyxl.Workbook()
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])

    ws = wb.create_sheet(title='2021', index=0)

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Inventory List'
    cell = ws['A1']
    cell.font = openpyxl.styles.Font(color='FF00FF', size=20)
    cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                  fgColor=openpyxl.styles.colors.Color(rgb='00FF00'))

    ws['A2'] = 'ID'
    ws['B2'] = 'Date in'
    ws['C2'] = 'Name'
    ws['D2'] = 'Quantity'
    ws['E2'] = 'Unit Price'
    ws['F2'] = 'Total'

    for col in range(ws.max_column):
        ws.column_dimensions[chr(ord('A') + col)].width = 15

    start_row = 3
    for (row, item) in enumerate(sample_items, start_row):
        ws.cell(row=row, column=1, value=row - 2)

        cell = ws.cell(row=row, column=2, value=item[0])
        cell.number_format = openpyxl.styles.numbers.FORMAT_DATE_YYYYMMDD2  # 'yyyy-mm-dd'

        ws.cell(row=row, column=3, value=item[1])

        cell = ws.cell(row=row, column=4, value=item[2])
        cell.number_format = openpyxl.styles.numbers.builtin_format_code(3)  # '#,##0'

        cell = ws.cell(row=row, column=5, value=item[3])
        cell.number_format = '#,##0'

        cell = ws.cell(row=row, column=6, value='= D{current_row} * E{current_row}'.format(current_row=cell.row))
        cell.number_format = openpyxl.styles.numbers.builtin_format_code(3)

    total_row = ws.max_row + 1
    if total_row > start_row:
        ws.cell(row=total_row, column=5, value='Total')
        cell = ws.cell(row=total_row, column=6, value='= SUM(F{}:F{})'.format(start_row, total_row - 1))
        cell.number_format = openpyxl.styles.numbers.builtin_format_code(3)

    wb.save(filepath)
    wb.close()


if __name__ == '__main__':
    write_to_file(os.getcwd() + '/test.xlsx')